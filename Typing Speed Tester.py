import tkinter as tk
import random
import time
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook


sentences = {
    "Easy": [
        "The sun is bright.",
        "I love my dog.",
        "It is a good day."
    ],
    "Medium": [
        "Python is a great programming language.",
        "Typing fast needs consistent effort.",
        "AI is changing the world rapidly."
    ],
    "Hard": [
        "Accuracy is more important than speed when learning to type.",
        "OpenAI's large language models are very powerful and flexible.",
        "Keyboard efficiency can dramatically impact coding productivity."
    ]
}

EXCEL_FILE = "high_scores.xlsx"
COUNTDOWN_TIME = 60  
class TypingSpeedTest:
    def __init__(self, root):
        self.root = root
        self.root.title("Typing Speed Tester")
        self.root.geometry("800x500")
        self.root.config(padx=20, pady=20)

        
        self.difficulty = tk.StringVar(value="Easy")
        self.sample_sentence = ""
        self.start_time = None
        self.remaining_time = COUNTDOWN_TIME
        self.timer_running = False
        self.timer_id = None

        self.create_widgets()
        self.set_dark_mode()
        self.set_sentence()
        self.load_high_score()

    def create_widgets(self):
        
        tk.Label(self.root, text="Select Difficulty:", font=("Arial", 12)).pack()
        options = tk.Frame(self.root)
        options.pack()
        for level in ["Easy", "Medium", "Hard"]:
            tk.Radiobutton(options, text=level, variable=self.difficulty, value=level, command=self.set_sentence).pack(side=tk.LEFT)

        
        self.title_label = tk.Label(self.root, text="Typing Speed Tester", font=("Arial", 20))
        self.title_label.pack(pady=10)

        
        self.text_label = tk.Label(self.root, text="", wraplength=700, font=("Arial", 14))
        self.text_label.pack(pady=10)

        
        self.text_entry = tk.Text(self.root, height=5, width=80, font=("Arial", 12))
        self.text_entry.pack(pady=10)
        self.text_entry.bind("<Key>", self.start_countdown)

        
        self.timer_label = tk.Label(self.root, text=f"Time Left: {COUNTDOWN_TIME} sec", font=("Arial", 12))
        self.timer_label.pack()

        
        self.submit_btn = tk.Button(self.root, text="Done", command=self.calculate_speed, font=("Arial", 12))
        self.submit_btn.pack(pady=10)

        
        self.result_label = tk.Label(self.root, text="", font=("Arial", 14))
        self.result_label.pack(pady=10)

        
        self.restart_btn = tk.Button(self.root, text="Try Again", command=self.reset_test, font=("Arial", 12))
        self.restart_btn.pack()

        
        self.high_score_label = tk.Label(self.root, text="", font=("Arial", 12))
        self.high_score_label.pack(pady=5)

    def set_dark_mode(self):
        dark_bg = "#1e1e1e"
        text_color = "#ffffff"
        highlight_color = "#B9375D"

        self.root.configure(bg=dark_bg)

        for widget in self.root.winfo_children():
            try:
                widget.configure(bg=dark_bg, fg=text_color)
            except:
                pass

        self.title_label.config(fg=highlight_color)
        self.text_label.config(fg=highlight_color)
        self.timer_label.config(fg="red")
        self.result_label.config(fg=text_color)
        self.high_score_label.config(fg=highlight_color)

    def set_sentence(self):
        level = self.difficulty.get()
        self.sample_sentence = random.choice(sentences[level])
        self.text_label.config(text=self.sample_sentence)

    def start_countdown(self, event):
        if not self.timer_running:
            self.start_time = time.time()
            self.timer_running = True
            self.countdown()

    def countdown(self):
        if self.remaining_time > 0:
            self.remaining_time -= 1
            self.timer_label.config(text=f"Time Left: {self.remaining_time} sec")
            self.timer_id = self.root.after(1000, self.countdown)
        else:
            self.calculate_speed()

    def calculate_speed(self):
        if self.timer_id:
            self.root.after_cancel(self.timer_id)

        typed_text = self.text_entry.get("1.0", tk.END).strip()
        total_words = len(typed_text.split())
        time_used = COUNTDOWN_TIME - self.remaining_time
        wpm = (total_words / time_used) * 60 if time_used > 0 else 0

        
        original_words = self.sample_sentence.split()
        typed_words = typed_text.split()
        correct_words = sum(1 for i in range(min(len(original_words), len(typed_words))) if original_words[i] == typed_words[i])
        accuracy = (correct_words / len(original_words)) * 100

        self.result_label.config(
            text=f"WPM: {int(wpm)} | Accuracy: {int(accuracy)}%"
        )

        self.save_high_score(int(wpm))
        self.text_entry.config(state="disabled")
        self.timer_running = False

    def reset_test(self):
        self.remaining_time = COUNTDOWN_TIME
        self.timer_label.config(text=f"Time Left: {COUNTDOWN_TIME} sec")
        self.start_time = None
        self.timer_running = False
        self.text_entry.config(state="normal")
        self.text_entry.delete("1.0", tk.END)
        self.result_label.config(text="")
        self.set_sentence()

    def save_high_score(self, score):
        level = self.difficulty.get()
        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        try:
            if not os.path.exists(EXCEL_FILE):
                wb = Workbook()
                ws = wb.active
                ws.title = "Scores"
                ws.append(["Date", "Difficulty", "WPM" , "Accuracy"])
            else:
                wb = load_workbook(EXCEL_FILE)
                ws = wb["Scores"]

            ws.append([current_date, level, score])
            wb.save(EXCEL_FILE)

        except Exception as e:
            print("Excel saving error:", e)

        self.load_high_score()

    def load_high_score(self):
        scores = {"Easy": 0, "Medium": 0, "Hard": 0}

        if os.path.exists(EXCEL_FILE):
            try:
                wb = load_workbook(EXCEL_FILE)
                ws = wb["Scores"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    _, level, wpm = row
                    if level in scores and int(wpm) > scores[level]:
                        scores[level] = int(wpm)
            except Exception as e:
                print("Excel read error:", e)

        high_text = "🏆 High Scores:\n" + "\n".join([f"{lvl}: {scores[lvl]} WPM" for lvl in scores])
        self.high_score_label.config(text=high_text)


if __name__ == "__main__":
    root = tk.Tk()
    app = TypingSpeedTest(root)
    root.mainloop()
